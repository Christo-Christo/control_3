import pandas as pd
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from control_3_trad import main as run_trad
from control_3_ul import main as run_ul

# Fungsi untuk membaca sheet filter dan mengembalikan daftar konfigurasi run
def load_filters(sheet_path, sheet_name):
    df = pd.read_excel(sheet_path, sheet_name=sheet_name)
    df = df[df.get("RUN").notna()]
    return df.to_dict(orient='records')

def process_trad_run(run_config):
    try:
        return run_config["RUN"], run_trad(run_config)
    except Exception as e:
        return run_config["RUN"], {"error": str(e)}

def process_ul_run(run_config):
    try:
        return run_config["RUN"], run_ul(run_config)
    except Exception as e:
        return run_config["RUN"], {"error": str(e)}
    
def write_to_excel_template(input_path, results_dict):
    wb = load_workbook(input_path, keep_vba=True)
    if "Control Template" in wb.sheetnames:
        template_sheet = wb["Control Template"]
    else:
        template_sheet = wb.active

    for run_name, result in results_dict.items():
        if "error" in result:
            continue

        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = str(run_name)[:31]

        # Penulisan Summary (misal di A2)
        if "summary_total" in result:
            summary_df = result["summary_total"]
            for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=2):
                for c_idx, value in enumerate(row, start=1):
                    new_sheet.cell(row=r_idx, column=c_idx, value=value)

        # Penulisan Tabel 2 (misal mulai A20)
        if "summary_tabel_2" in result:
            df2 = result["summary_tabel_2"]
            for r_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=True), start=20):
                for c_idx, value in enumerate(row, start=1):
                    new_sheet.cell(row=r_idx, column=c_idx, value=value)

        # Penulisan Tabel 3 (misal mulai A50)
        if "summary_tabel_3" in result:
            df3 = result["summary_tabel_3"]
            for r_idx, row in enumerate(dataframe_to_rows(df3, index=False, header=True), start=50):
                for c_idx, value in enumerate(row, start=1):
                    new_sheet.cell(row=r_idx, column=c_idx, value=value)

        # Tambah tabel lainnya jika ada

    wb.save(input_path)


def main(INPUT_EXCEL_PATH):
    trad_filters = load_filters(INPUT_EXCEL_PATH, "FILTER_TRAD")
    ul_filters = load_filters(INPUT_EXCEL_PATH, "FILTER_UL")

    trad_results = {}
    ul_results = {}

    with ThreadPoolExecutor() as executor:
        trad_futures = {executor.submit(process_trad_run, cfg): cfg["RUN"] for cfg in trad_filters}
        ul_futures = {executor.submit(process_ul_run, cfg): cfg["RUN"] for cfg in ul_filters}

        for future in as_completed(trad_futures):
            run_name, result = future.result()
            trad_results[run_name] = result

        for future in as_completed(ul_futures):
            run_name, result = future.result()
            ul_results[run_name] = result

    all_results = {**trad_results, **ul_results}
    write_to_excel_template(INPUT_EXCEL_PATH, all_results)

