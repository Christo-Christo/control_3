import pandas as pd
import glob
import os
from concurrent.futures import ProcessPoolExecutor
import re
from openpyxl import load_workbook
from itertools import zip_longest
import traceback

columns_to_sum_argo = [
    'prm_inc','lrc_cl_ins','lrc_cl_inv','r_exp_m','r_acq_cost',
    'cov_units','dac_cov_units','dac','nattr_exp_acq','nattr_exp_inv',
    'nattr_exp_maint','nattr_exp','c_sar','pv_r_exp_m','pv_surr',
    'lrc_cl_ins_dth','lrc_cl_inv_dth','lrc_cl_inv_surr','lrc_cl_inv_mat',
    'lrc_cl_inv_ann'
]

columns_to_sum_rafm = [
    'period', 'prm_inc', 'lrc_cl_ins', 'lrc_cl_inv', 'r_exp_m', 'r_acq_cost',
    'nattr_exp_acq', 'nattr_exp_inv', 'nattr_exp_maint',
    'lrc_cl_ins_dth', 'lrc_cl_inv_dth', 'lrc_cl_inv_surr', 'lrc_cl_inv_mat',
    'lrc_cl_inv_ann'
]

columns_to_sum_uvsg = [
    'period', 'prm_inc', 'lrc_cl_ins', 'lrc_cl_inv', 'r_exp_m', 'r_acq_cost',
    'nattr_exp_acq', 'nattr_exp_inv', 'nattr_exp_maint',
    'lrc_cl_ins_dth', 'lrc_cl_inv_dth', 'lrc_cl_inv_surr', 'lrc_cl_inv_mat',
    'lrc_cl_inv_ann'
]

cols_to_compare = [
    'prm_inc','lrc_cl_ins','lrc_cl_inv','r_exp_m','r_acq_cost',
    'cov_units','dac_cov_units','dac','nattr_exp_acq','nattr_exp_inv',
    'nattr_exp_maint','nattr_exp','c_sar','pv_r_exp_m','pv_surr',
    'lrc_cl_ins_dth','lrc_cl_inv_dth','lrc_cl_inv_surr','lrc_cl_inv_mat',
    'lrc_cl_inv_ann'
]

additional_columns = ['cov_units', 'pv_r_exp_m', 'pv_surr']
c_sar = ['c_sar']
additional_columns_uvsg = ['cov_units', 'pv_r_exp_m', 'pv_surr']
u_sar = ['u_sar']
target_sheets = ['extraction_IDR', 'extraction_USD']
global_filter_rafm = None
global_filter_uvsg = None


def parse_numeric_fast(val):
    """Fast numeric parser optimized for Excel data"""
    if val is None or val == '':
        return None
    
    if isinstance(val, (int, float)):
        return float(val)
    
    if isinstance(val, str):
        s = val.strip()
        
        if not s or s.lower() in ['none', 'nan', 'n/a', '-']:
            return None
        
        try:
            return float(s)
        except ValueError:
            pass
        
        try:
            s = s.replace('\xa0', '').replace(' ', '')
            
            is_percent = s.endswith('%')
            if is_percent:
                s = s[:-1]
            
            is_negative = False
            if s.startswith('(') and s.endswith(')'):
                is_negative = True
                s = s[1:-1]
            
            comma_count = s.count(',')
            dot_count = s.count('.')
            
            if comma_count == 0 and dot_count <= 1:
                result = float(s)
            elif dot_count == 0 and comma_count > 0:
                result = float(s.replace(',', ''))
            elif comma_count > 0 and dot_count > 0:
                last_comma = s.rfind(',')
                last_dot = s.rfind('.')
                
                if last_dot > last_comma:
                    result = float(s.replace(',', ''))
                else:
                    result = float(s.replace('.', '').replace(',', '.'))
            elif dot_count > 1:
                result = float(s.replace('.', ''))
            elif comma_count == 1:
                comma_pos = s.find(',')
                digits_after = len(s) - comma_pos - 1
                
                if digits_after == 2 and comma_pos <= 3:
                    result = float(s.replace(',', '.'))
                else:
                    result = float(s.replace(',', ''))
            else:
                result = float(s)
            
            if is_negative:
                result = -result
            if is_percent:
                result = result / 100.0
            
            return result
            
        except (ValueError, AttributeError):
            return None
    
    try:
        return float(val)
    except:
        return None


def process_argo_file(file_path):
    """Optimized ARGO file processing"""
    file_name_argo = os.path.splitext(os.path.basename(file_path))[0]
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        sheet = wb['Sheet1']
        
        data = list(sheet.values)
        if not data:
            wb.close()
            return {'File_Name': file_name_argo}
        
        header = data[0]
        col_index = {col: i for i, col in enumerate(header) if col in columns_to_sum_argo}
        sums = {col: 0 for col in col_index}
        
        for row in data[1:]:
            for col, idx in col_index.items():
                if idx < len(row):
                    parsed_val = parse_numeric_fast(row[idx])
                    if parsed_val is not None:
                        sums[col] += parsed_val
        wb.close()
    except Exception as e:
        print(f"❌ Gagal proses {file_name_argo}: {e}")
        sums = {}
    
    sums['File_Name'] = file_name_argo
    return sums


def try_match_filter(filter_df, file_name):
    """Optimized filter matching"""
    base = os.path.splitext(file_name)[0]
    candidates = [
        file_name, base, file_name.lower(), base.lower(),
        base.replace('_',' ').lower(), base.replace(' ','_').lower()
    ]
    
    for c in candidates:
        m = filter_df[filter_df['File Name'].str.lower() == str(c).lower()]
        if not m.empty:
            return m
    
    for c in candidates:
        m = filter_df[filter_df['File Name'].str.lower().str.contains(str(c).lower(), na=False)]
        if not m.empty:
            return m
    
    return filter_df[filter_df['File Name'] == file_name]


def process_rafm_file(args):
    """Optimized RAFM file processing"""
    file_path, file_name, filter_df = args
    try:
        match = try_match_filter(filter_df, file_name)
        if match.empty:
            print(f"⚠️ Filter match empty for RAFM file: {file_name}")
            return None

        total_sums = {col: 0.0 for col in columns_to_sum_rafm}
        additional_sums = {col: 0.0 for col in additional_columns}
        csar_columns = {col: 0.0 for col in c_sar}

        speed = int(match['Speed Duration'].values[0])
        exclude = str(match['Exclude Year'].values[0])
        include = str(match['Include Year'].values[0])
        sar = int(match['C_sar'].values[0])

        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        
        for sheet_name in target_sheets:
            try:
                actual_sheetnames = [s.strip() for s in wb.sheetnames]
                matched_sheet = None
                for s in actual_sheetnames:
                    if s.lower().strip() == sheet_name.lower().strip():
                        matched_sheet = s
                        break
                
                if matched_sheet is None:
                    continue

                sheet = wb[matched_sheet]
                data = list(sheet.values)
                if not data:
                    continue
                
                header = data[0]
                expected = [c.lower() for c in columns_to_sum_rafm + additional_columns + c_sar]
                col_index = {}
                
                for i, col in enumerate(header):
                    col_name = str(col).strip().lower() if col is not None else ''
                    if col_name in expected or col_name == 'goc' or col_name == 'period':
                        col_index[col_name] = i

                if 'goc' not in col_index:
                    continue

                for row in data[1:]:
                    val_goc = ''
                    try:
                        idx_goc = col_index.get('goc')
                        if idx_goc is not None and idx_goc < len(row):
                            val_goc = str(row[idx_goc]) if row[idx_goc] is not None else ''
                    except Exception:
                        val_goc = ''

                    period_idx = col_index.get('period')
                    period_value = None
                    if period_idx is not None and period_idx < len(row):
                        period_raw = row[period_idx]
                        period_value = parse_numeric_fast(period_raw)
                        if period_value is not None:
                            period_value = int(period_value)

                    skip_row = False
                    if include != '-' and exclude != '-':
                        if include not in val_goc or exclude in val_goc:
                            skip_row = True
                    elif include != '-':
                        if include not in val_goc:
                            skip_row = True
                    elif exclude != '-':
                        if exclude in val_goc:
                            skip_row = True
                    if skip_row:
                        continue

                    if period_value is not None and period_value > speed:
                        for col in columns_to_sum_rafm:
                            idx = col_index.get(col.lower())
                            if idx is not None and idx < len(row):
                                v = parse_numeric_fast(row[idx])
                                if v is not None and v != 0:
                                    total_sums[col] += v

                    if period_value is not None and period_value >= 0:
                        for col in additional_columns:
                            idx = col_index.get(col.lower())
                            if idx is not None and idx < len(row):
                                v = parse_numeric_fast(row[idx])
                                if v is not None and v != 0:
                                    additional_sums[col] += v

                    if period_value is not None and period_value >= sar:
                        for col in c_sar:
                            idx = col_index.get(col.lower())
                            if idx is not None and idx < len(row):
                                v = parse_numeric_fast(row[idx])
                                if v is not None and v != 0:
                                    csar_columns[col] += v

            except Exception:
                print(f"Error processing sheet {sheet_name} in file {file_name}:")
                traceback.print_exc()
                continue

        wb.close()
        total_sums['File_Name'] = file_name
        additional_sums['File_Name'] = file_name
        csar_columns['File_Name'] = file_name

        return total_sums, additional_sums, csar_columns

    except Exception:
        print(f"Fatal error in processing file {file_name}:")
        traceback.print_exc()
        return None


def process_uvsg_file(args):
    """Optimized UVSG file processing"""
    file_path, file_name, filter_df = args

    match = try_match_filter(filter_df, file_name)
    if match.empty:
        print(f"⚠️ File UVSG '{file_name}' tidak berhasil diproses atau hasilnya None.")
        return None

    match = match.iloc[0]

    try:
        speed = int(match['Speed Duration'])
        exclude = str(match['Exclude Year'])
        include = str(match['Include Year'])
        sar = int(match['C_sar'])
    except Exception as e:
        print(f"❌ Error membaca filter UVSG untuk {file_name}: {e}")
        return None

    total_sums = {col: 0.0 for col in columns_to_sum_uvsg}
    additional_sums = {col: 0.0 for col in additional_columns_uvsg}
    usar_columns = {col: 0.0 for col in u_sar}

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        
        for sheet_name in target_sheets:
            try:
                actual_sheets = [s.strip() for s in wb.sheetnames]
                matched_sheet = None
                for s in actual_sheets:
                    if s.lower() == sheet_name.lower().strip():
                        matched_sheet = s
                        break
                if matched_sheet is None:
                    continue

                sheet = wb[matched_sheet]
                data = list(sheet.values)
                if not data:
                    continue

                header = data[0]
                header_lower = [str(h).strip().lower() if h else '' for h in header]
                col_index = {}
                
                for col in columns_to_sum_uvsg + additional_columns_uvsg + u_sar + ['goc', 'period']:
                    cl = col.lower()
                    if cl in header_lower:
                        col_index[cl] = header_lower.index(cl)

                if 'goc' not in col_index:
                    continue

                for row in data[1:]:
                    val_goc = str(row[col_index['goc']]) if col_index['goc'] < len(row) and row[col_index['goc']] is not None else ''
                    period_idx = col_index.get('period')
                    period_value = None
                    if period_idx is not None and period_idx < len(row):
                        period_value = parse_numeric_fast(row[period_idx])
                        if period_value is not None:
                            period_value = int(period_value)

                    skip_row = False
                    if include != '-' and exclude != '-':
                        if include not in val_goc or exclude in val_goc:
                            skip_row = True
                    elif include != '-':
                        if include not in val_goc:
                            skip_row = True
                    elif exclude != '-':
                        if exclude in val_goc:
                            skip_row = True
                    if skip_row:
                        continue

                    if period_value is not None and period_value > speed:
                        for col in columns_to_sum_uvsg:
                            idx = col_index.get(col.lower())
                            if idx is not None and idx < len(row):
                                v = parse_numeric_fast(row[idx])
                                if v is not None and v != 0:
                                    total_sums[col] += v

                    if period_value is not None and period_value >= 0:
                        for col in additional_columns_uvsg:
                            idx = col_index.get(col.lower())
                            if idx is not None and idx < len(row):
                                v = parse_numeric_fast(row[idx])
                                if v is not None and v != 0:
                                    additional_sums[col] += v

                    if period_value is not None and period_value >= sar:
                        for col in u_sar:
                            idx = col_index.get(col.lower())
                            if idx is not None and idx < len(row):
                                v = parse_numeric_fast(row[idx])
                                if v is not None and v != 0:
                                    usar_columns[col] += v
            except Exception:
                print(f"Error processing sheet {sheet_name} in file {file_name}:")
                traceback.print_exc()
                continue
                
        wb.close()
        total_sums['File_Name'] = file_name
        additional_sums['File_Name'] = file_name
        usar_columns['File_Name'] = file_name

        return total_sums, additional_sums, usar_columns

    except Exception as e:
        print(f"❌ Gagal membaca file UVSG {file_name}: {e}")
        return None


def main(params):
    """Optimized main function"""
    global global_filter_rafm, global_filter_uvsg

    input_excel = params['input excel']

    excel_file = pd.ExcelFile(input_excel)
    code = pd.read_excel(excel_file, sheet_name='Code')
    sign_logic = pd.read_excel(excel_file, sheet_name='Sign Logic')
    control = pd.read_excel(excel_file, sheet_name='Control')
    file_path_df = pd.read_excel(excel_file, sheet_name='File Path')
    global_filter_rafm = pd.read_excel(excel_file, sheet_name='Filter RAFM')
    global_filter_uvsg = pd.read_excel(excel_file, sheet_name='Filter UVSG')
    excel_file.close()

    path_map = dict(zip(file_path_df['Name'].str.lower(), file_path_df['File Path']))
    folder_path_argo = path_map.get('argo', '')
    folder_path_rafm = path_map.get('rafm', '')
    folder_path_uvsg = path_map.get('uvsg', '')
    rafm_manual_path = path_map.get('rafm manual', '')

    file_paths_argo = [f for f in glob.glob(os.path.join(folder_path_argo, '*.xlsx')) 
                       if not os.path.basename(f).startswith('~')]
    
    optimal_workers = min(os.cpu_count() or 4, max(len(file_paths_argo), 1))
    
    with ProcessPoolExecutor(max_workers=optimal_workers) as executor:
        summary_rows_argo = list(executor.map(process_argo_file, file_paths_argo))
    
    cf_argo = pd.DataFrame(summary_rows_argo)
    if 'File_Name' in cf_argo.columns:
        cols = ['File_Name'] + [col for col in cf_argo.columns if col != 'File_Name']
        cf_argo = cf_argo[cols]
    
    cf_argo = cf_argo.rename(columns={'File_Name': 'ARGO File Name'})
    
    mask = code['RAFM File Name'].astype(str).str.contains('_ori', regex=True, na=False)
    code = code[~mask].copy()
    
    cf_argo = pd.merge(code, cf_argo, on='ARGO File Name', how='left')
    
    columns_to_drop = [col for col in ['RAFM File Name', 'UVSG File Name'] if col in cf_argo.columns]
    if columns_to_drop:
        cf_argo = cf_argo.drop(columns=columns_to_drop)
    
    file_paths_rafm = [f for f in glob.glob(os.path.join(folder_path_rafm, '*.xlsx')) 
                       if not os.path.basename(f).startswith('~')]
    file_entries = [(f, os.path.splitext(os.path.basename(f))[0], global_filter_rafm) 
                    for f in file_paths_rafm]

    with ProcessPoolExecutor(max_workers=optimal_workers) as executor:
        results = list(executor.map(process_rafm_file, file_entries))
    
    summary_rows_rafm = []
    additional_summary_rows = []
    csar_summary = []
    for result in results:
        if result:
            total_sums, additional_sums, csar_columns = result
            summary_rows_rafm.append(total_sums)
            additional_summary_rows.append(additional_sums)
            csar_summary.append(csar_columns)

    combined_summary = []
    for main_row, add_row, csar_row in zip_longest(summary_rows_rafm, additional_summary_rows, csar_summary):
        combined_row = {**main_row, **add_row, **csar_row}
        combined_summary.append(combined_row)

    all_runs = ['11', '21', '31', '41']
    pattern = '|'.join([f'run{r}' for r in all_runs])
    combined_summary = pd.DataFrame(combined_summary)
    
    def add_ori_if_run(x):
        for r in all_runs:
            if re.search(fr'run_?{r}', x, re.IGNORECASE):
                return x + "_ori"
        return x
    
    combined_summary['File_Name'] = combined_summary['File_Name'].apply(add_ori_if_run)
    cf_rafm_1 = combined_summary.copy()
    cols = ['File_Name'] + [col for col in cf_rafm_1.columns if col != 'File_Name']
    cf_rafm_1 = cf_rafm_1[cols]

    code_rafm = code.copy()
    if 'UVSG File Name' in code_rafm.columns:
        code_rafm = code_rafm.drop(columns=['UVSG File Name'])
    
    cf_rafm = cf_rafm_1.rename(columns={'File_Name': 'RAFM File Name'})
    run1_ori = cf_rafm[cf_rafm['RAFM File Name'].str.contains(pattern, case=False, na=False)]
    cf_rafm_merge = pd.merge(code_rafm, cf_rafm, on="RAFM File Name", how="left")
    cf_rafm_merge.fillna(0, inplace=True)

    numeric_cols = cf_rafm_merge.select_dtypes(include='number').columns
    sum_rows = cf_rafm_merge[cf_rafm_merge['RAFM File Name'].str.contains("SUM_", na=False)]

    for idx, row in sum_rows.iterrows():
        rafm_value = row['RAFM File Name']

        if 'SUM_' in rafm_value:
            keyword = rafm_value.split('SUM_')[-1]
            pattern_search = re.escape(keyword).replace("-", "[-_]?")
            matched_rows = cf_rafm_merge[cf_rafm_merge['ARGO File Name'].str.contains(
                pattern_search, case=False, regex=True, na=False)]

            total_values = matched_rows[numeric_cols].sum()

            for col in numeric_cols:
                cf_rafm_merge.at[idx, col] = total_values[col]

    columns_to_drop = [col for col in ['ARGO File Name', 'period'] if col in cf_rafm_merge.columns]
    if columns_to_drop:
        cf_rafm = cf_rafm_merge.drop(columns=columns_to_drop)
    else:
        cf_rafm = cf_rafm_merge.copy()
    
    cf_rafm = pd.concat([cf_rafm, run1_ori], ignore_index=True)
    if 'period' in cf_rafm.columns:
        cf_rafm = cf_rafm.drop(columns=['period'])
    
    cf_rafm['dac_cov_units'] = cf_rafm['cov_units']
    cf_rafm['dac'] = -cf_rafm['r_acq_cost']
    
    nattr_exp = ['nattr_exp_acq', 'nattr_exp_inv', 'nattr_exp_maint']
    for col in nattr_exp:
        cf_rafm[col] = cf_rafm[col].astype(str).str.replace(',', '').astype(float)
    cf_rafm['nattr_exp'] = cf_rafm['nattr_exp_acq'] + cf_rafm['nattr_exp_inv'] + cf_rafm['nattr_exp_maint']
    
    file_paths_uvsg = [f for f in glob.glob(os.path.join(folder_path_uvsg, '*.xlsx')) 
                       if not os.path.basename(f).startswith('~')]

    summary_rows_uvsg = []
    additional_summary_rows = []
    usar_summary_uvsg = []

    if file_paths_uvsg:
        try:
            file_entries = [(f, os.path.splitext(os.path.basename(f))[0], global_filter_uvsg) 
                           for f in file_paths_uvsg]

            with ProcessPoolExecutor(max_workers=optimal_workers) as executor:
                results_uvsg = list(executor.map(process_uvsg_file, file_entries))

            for entry, result in zip(file_entries, results_uvsg):
                if isinstance(result, tuple) and len(result) == 3:
                    total_sums, additional_sums, usar_columns = result
                    summary_rows_uvsg.append(total_sums)
                    additional_summary_rows.append(additional_sums)
                    usar_summary_uvsg.append(usar_columns)

        except Exception as e:
            print(f"❌ Terjadi kesalahan saat memproses file UVSG: {e}")

    combined_summary = []
    for main_row, add_row, usar_row in zip_longest(summary_rows_uvsg, additional_summary_rows, usar_summary_uvsg):
        combined_row = {**main_row, **add_row, **usar_row}
        combined_summary.append(combined_row)

    if combined_summary:
        uvsg_1 = pd.DataFrame(combined_summary)
        if 'File_Name' in uvsg_1.columns:
            cols = ['File_Name'] + [col for col in uvsg_1.columns if col != 'File_Name']
            uvsg_1 = uvsg_1[cols]
        else:
            uvsg_1 = pd.DataFrame(columns=['File_Name'] + columns_to_sum_uvsg + additional_columns_uvsg + u_sar)
    else:
        uvsg_1 = pd.DataFrame(columns=['File_Name'] + columns_to_sum_uvsg + additional_columns_uvsg + u_sar)

    uvsg_1 = uvsg_1.rename(columns={'u_sar': 'c_sar'})
    if 'period' in uvsg_1.columns:
        uvsg_1 = uvsg_1.drop(columns=['period'])
    
    uvsg_1['dac_cov_units'] = uvsg_1['cov_units']
    uvsg_1['dac'] = -uvsg_1['r_acq_cost']
    for col in nattr_exp:
        uvsg_1[col] = uvsg_1[col].astype(str).str.replace(',', '').astype(float)
    uvsg_1['nattr_exp'] = uvsg_1['nattr_exp_acq'] + uvsg_1['nattr_exp_inv'] + uvsg_1['nattr_exp_maint']

    uvsg_2 = uvsg_1.copy()
    code_uvsg = code.copy()
    if 'ARGO File Name' in code_uvsg.columns:
        code_uvsg = code_uvsg.drop(columns=['ARGO File Name'])
    
    uvsg = uvsg_2.rename(columns={'File_Name': 'UVSG File Name'})
    uvsg_merged = pd.merge(code_uvsg, uvsg, on="UVSG File Name", how="left")
    uvsg_merged.fillna(0, inplace=True)
    if 'RAFM File Name' in uvsg_merged.columns:
        uvsg = uvsg_merged.drop(columns=['RAFM File Name'])
    else:
        uvsg = uvsg_merged.copy()
    
    rafm_manual = pd.read_excel(rafm_manual_path, sheet_name='Sheet1', engine='openpyxl')
    rafm_manual = rafm_manual.drop(columns=['No','Update (Y/N)','Shift Dur','Cohort','C_sar'])
    rafm_manual = rafm_manual.fillna(0)

    final = code.copy()
    for col in cols_to_compare:
        if col not in code.columns:
            final[col] = pd.NA
    logic_row = sign_logic.iloc[0]

    mapping_code = global_filter_rafm.drop(columns={'File Name'})
    mapping = pd.concat([code_rafm, mapping_code], axis=1)
    mask = mapping['RAFM File Name'].astype(str).str.contains('_ori', regex=True, na=False)
    mapping = mapping[~mask].copy()
    
    cf_rafm = cf_rafm.groupby('RAFM File Name', as_index=False).first()
    global_filter_rafm = global_filter_rafm.rename(columns={'File Name':'RAFM File Name'})
    cf_rafm = pd.merge(global_filter_rafm, cf_rafm, on='RAFM File Name', how='left')
    
    global_filter_uvsg = global_filter_uvsg.groupby('File Name', as_index=False).first()
    global_filter_uvsg = global_filter_uvsg.rename(columns={'File Name':'UVSG File Name'})
    uvsg = pd.merge(uvsg, global_filter_uvsg, on='UVSG File Name', how='left')

    valid_cols = [col for col in logic_row.index if col in cf_argo.columns]
    
    def check_sign(val, logic_sign):
        if pd.isna(val):
            return 0
        if logic_sign == 1:
            return 1 if val < 0 else 0
        elif logic_sign == "-":
            return 0  
        elif logic_sign == -1:
            return 1 if val > 0 else 0 
        return 0 

    check_sign_summary_row = {
        col: cf_argo[col].apply(lambda val: check_sign(val, logic_row[col])).sum()
        for col in valid_cols
    }

    for col in cf_argo.columns:
        if col not in check_sign_summary_row:
            check_sign_summary_row[col] = None
    
    check_sign_summary = pd.DataFrame([check_sign_summary_row])[cf_argo.columns]
    cf_argo = pd.concat([cf_argo, check_sign_summary], ignore_index=True)
    check_sign_total = sum(val for val in check_sign_summary_row.values() if isinstance(val, (int, float)))
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = check_sign_total
    
    index_labels = list(range(1, len(cf_argo))) + ['check sign']
    cf_argo.insert(0, 'No', index_labels)
    cf_argo = pd.concat([cf_argo, sign_logic], ignore_index=True)
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = 'Sign Logic'
    
    index_labels_manual = list(range(1, len(rafm_manual)+1))
    rafm_manual.insert(0, 'No', index_labels_manual)
    index_labels_final = list(range(1, len(final)+1))
    final.insert(0, 'No', index_labels_final)

    control['check sign'] = ''
    control['result'] = ''

    val_year_idx = control[control.iloc[:, 0] == 'Val Year'].index
    if not val_year_idx.empty:
        idx = val_year_idx[0]
        control.at[idx, 'check sign'] = 'Check Sign'
        control.at[idx, 'result'] = check_sign_total

    if 'UVSG File Name' in uvsg.columns:
        last_3_cols = uvsg.columns[-4:].tolist()
        other_cols_uvsg = [col for col in uvsg.columns if col not in last_3_cols and col != 'UVSG File Name']
        uvsg = uvsg[['UVSG File Name'] + last_3_cols + other_cols_uvsg]
    
    index_labels_rafm = list(range(1, len(cf_rafm)+1))
    cf_rafm.insert(0, 'No', index_labels_rafm)
    index_labels_uvsg = list(range(1, len(uvsg)+1))
    uvsg.insert(0, 'No', index_labels_uvsg)
    
    return {
        'Control': control,
        'Code': mapping,
        "CF ARGO AZTRAD": cf_argo,
        "RAFM Output AZTRAD": cf_rafm,
        "RAFM Output Manual": rafm_manual,
        "RAFM Output AZUL_PI": uvsg,
        "Checking Summary AZTRAD": final
    }


if __name__ == '__main__':
    import multiprocessing
    multiprocessing.freeze_support()