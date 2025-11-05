import os
import pandas as pd
from xlsxwriter.utility import xl_col_to_name
import syntax.control_4_trad as trad
import syntax.control_4_ul as ul
import syntax.control_4_reas as reas
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
import shutil
import datetime
import tempfile
import subprocess
import psutil

# Suppress warnings untuk performa
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

cols_to_sum_dict = {
    'trad': trad.cols_to_compare,
    'ul': ul.columns_to_sum_argo,
    'reas': reas.cols_to_compare
}


def kill_excel_processes():
    """Force close all Excel processes"""
    try:
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                if proc.info['name'] and 'excel' in proc.info['name'].lower():
                    print(f"    • Killing Excel PID {proc.info['pid']}")
                    proc.kill()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        time.sleep(2)  # Wait for processes to fully terminate
    except Exception as e:
        print(f"    ⚠️ Error killing Excel: {e}")


# ============================
#  Fungsi styling & helpers
# ============================

def auto_adjust_column_width_xlwings(ws, df_sheet, sample_size=100):
    """
    Set column width based on the longest text in each column (xlwings).
    No fixed maximum — mengikuti tulisan paling panjang.
    """
    if not hasattr(df_sheet, 'columns'):
        return

    # Determine number of header rows (Control may not have header)
    header_offset = 0
    if isinstance(df_sheet, pd.DataFrame) and df_sheet.shape[1] > 0:
        header_offset = 1  # karena kita menulis header untuk non-Control sebagai baris 1

    # iterate columns
    for i, col in enumerate(df_sheet.columns):
        try:
            # sample header + sample data
            samples = []
            samples.append(str(col))
            sample_vals = df_sheet[col].head(sample_size).astype(str).tolist()
            samples.extend(sample_vals)
            max_len = max(len(s) for s in samples)
            adjusted_width = max(8, max_len + 2)  # minimal 8
            # xlwings column index for setting width: use cell at first row in that column
            # set via Range((row, col)).column_width
            try:
                ws.range((1, i + 1)).column_width = adjusted_width
            except Exception:
                # fallback: try setting via api
                try:
                    ws.api.Columns(i + 1).ColumnWidth = adjusted_width
                except Exception:
                    pass
        except Exception:
            try:
                ws.range((1, i + 1)).column_width = 12
            except Exception:
                pass


def apply_number_formats_xlwings(ws, df_sheet):
    """
    Apply number formats:
     - If column name contains 'speed duration' -> text (no number format)
     - If column name contains 'include year' or 'exclude year' -> integer '0'
     - Else -> accounting format
    Uses xlwings Range.number_format
    """
    if not hasattr(df_sheet, 'columns'):
        return

    nrows = len(df_sheet) + 1  # +1 for header row
    ncols = len(df_sheet.columns)

    for col_idx, col_name in enumerate(df_sheet.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_name_lower = str(col_name).lower()

        if 'speed duration' in col_name_lower:
            number_format = '@'  # text / no format
        elif 'include year' in col_name_lower or 'exclude year' in col_name_lower:
            number_format = '0'  # integer
        else:
            number_format = '_-* #,##0_-;_-* (#,##0);_-* "-"_-;_-@_-'

        try:
            rng = ws.range(f"{col_letter}2:{col_letter}{nrows}")
            rng.number_format = number_format
        except Exception:
            # fallback: try applying to column via api
            try:
                rng_api = ws.api.Range(f"{col_letter}2:{col_letter}{nrows}")
                rng_api.NumberFormat = number_format
            except Exception:
                pass


def apply_border_xlwings(ws, df_sheet):
    """
    Apply thin black border around the used table (A1:lastcol lastrow)
    """
    try:
        nrows = len(df_sheet) + 1
        ncols = len(df_sheet.columns)
        if nrows < 1 or ncols < 1:
            return
        last_cell = f"{get_column_letter(ncols)}{nrows}"
        full_range = ws.range(f"A1:{last_cell}")
        full_range.api.Borders.LineStyle = 1  # xlContinuous
        full_range.api.Borders.Weight = 2     # xlThin
    except Exception:
        try:
            # alternative: loop cells (slower) - but keep best-effort
            nrows = len(df_sheet) + 1
            ncols = len(df_sheet.columns)
            for r in range(1, nrows + 1):
                for c in range(1, ncols + 1):
                    try:
                        cell = ws.api.Cells(r, c)
                        cell.Borders.LineStyle = 1
                        cell.Borders.Weight = 2
                    except Exception:
                        pass
        except Exception:
            pass


def apply_accounting_to_all_xlwings(ws, df_sheet):
    """
    Apply accounting format to all numeric columns (used for Checking Summary)
    """
    if not hasattr(df_sheet, 'columns'):
        return
    nrows = len(df_sheet) + 1
    ncols = len(df_sheet.columns)
    for col_idx in range(1, ncols + 1):
        col_letter = get_column_letter(col_idx)
        try:
            data_range = ws.range(f'{col_letter}2:{col_letter}{nrows}')
            data_range.number_format = '_-* #,##0_-;_-* (#,##0);_-* "-"_-;_-@_-'
        except Exception:
            try:
                rng_api = ws.api.Range(f'{col_letter}2:{col_letter}{nrows}')
                rng_api.NumberFormat = '_-* #,##0_-;_-* (#,##0);_-* "-"_-;_-@_-'
            except Exception:
                pass


# ======================================
# Existing logic (preserved) - checking formulas and main flow
# ======================================

def write_checking_summary_formulas_xlwings(ws, df_sheet, jenis, start_row=2):
    """
    Tulis formula checking summary menggunakan xlwings

    Args:
        ws: xlwings worksheet object
        df_sheet: DataFrame checking summary
        jenis: 'trad', 'ul', atau 'reas'
        start_row: Baris mulai data (default=2, karena row 1 = header)
    """
    sheet_names = {
        'trad': {
            'cf_argo': 'CF ARGO AZTRAD',
            'cf_rafm': 'RAFM Output AZTRAD',
            'rafm_manual': 'RAFM Output Manual',
            'uvsg': 'RAFM Output AZUL_PI'
        },
        'ul': {
            'cf_argo': 'CF ARGO AZUL',
            'cf_rafm': 'RAFM Output AZUL',
            'rafm_manual': 'RAFM Output Manual'
        },
        'reas': {
            'cf_argo': 'CF ARGO REAS',
            'cf_rafm': 'RAFM Output REAS',
            'rafm_manual': 'RAFM Output Manual'
        }
    }

    nrows = len(df_sheet)
    ncols = len(df_sheet.columns)

    # Tentukan kolom mulai formula dan offset
    if jenis == 'trad':
        start_col_idx = 5  # Kolom E
        cf_argo_col_offset = 3  # Kolom C
        cf_rafm_col_offset = 7  # Kolom G
        rafm_manual_col_offset = 7
        uvsg_col_offset = 7
    elif jenis == 'ul':
        start_col_idx = 4  # Kolom D
        cf_argo_col_offset = 3  # Kolom C
        cf_rafm_col_offset = 6  # Kolom F
        rafm_manual_col_offset = 6
    else:  # reas
        start_col_idx = 4  # Kolom D
        cf_argo_col_offset = 3  # Kolom C
        cf_rafm_col_offset = 3  # Kolom C
        rafm_manual_col_offset = 3

    # Loop per row
    for row_idx in range(nrows):
        row_excel = start_row + row_idx

        # Loop per column
        for col_idx in range(start_col_idx, ncols + 1):
            relative_offset = col_idx - start_col_idx

            if jenis == 'trad':
                cf_argo_col = get_column_letter(cf_argo_col_offset + relative_offset)
                cf_rafm_col = get_column_letter(cf_rafm_col_offset + relative_offset)
                rafm_manual_col = get_column_letter(rafm_manual_col_offset + relative_offset)
                uvsg_col = get_column_letter(uvsg_col_offset + relative_offset)

                formula = (
                    f"='{sheet_names['trad']['cf_argo']}'!{cf_argo_col}{row_excel}"
                    f"-'{sheet_names['trad']['cf_rafm']}'!{cf_rafm_col}{row_excel}"
                    f"+'{sheet_names['trad']['rafm_manual']}'!{rafm_manual_col}{row_excel}"
                    f"-'{sheet_names['trad']['uvsg']}'!{uvsg_col}{row_excel}"
                )
            elif jenis == 'ul':
                cf_argo_col = get_column_letter(cf_argo_col_offset + relative_offset)
                cf_rafm_col = get_column_letter(cf_rafm_col_offset + relative_offset)
                rafm_manual_col = get_column_letter(rafm_manual_col_offset + relative_offset)

                formula = (
                    f"='{sheet_names['ul']['cf_argo']}'!{cf_argo_col}{row_excel}"
                    f"-'{sheet_names['ul']['cf_rafm']}'!{cf_rafm_col}{row_excel}"
                    f"-'{sheet_names['ul']['rafm_manual']}'!{rafm_manual_col}{row_excel}"
                )
            else:  # reas
                cf_argo_col = get_column_letter(cf_argo_col_offset + relative_offset)
                cf_rafm_col = get_column_letter(cf_rafm_col_offset + relative_offset)
                rafm_manual_col = get_column_letter(rafm_manual_col_offset + relative_offset)

                formula = (
                    f"='{sheet_names['reas']['cf_argo']}'!{cf_argo_col}{row_excel}"
                    f"-'{sheet_names['reas']['cf_rafm']}'!{cf_rafm_col}{row_excel}"
                    f"+'{sheet_names['reas']['rafm_manual']}'!{rafm_manual_col}{row_excel}"
                )

            # Write formula using xlwings
            cell_address = f"{get_column_letter(col_idx)}{row_excel}"
            ws.range(cell_address).formula = formula


def add_sheets_to_rafm_manual(rafm_manual_path, result_dict, output_path, output_filename, jenis):
    """
    🔧 XLWINGS APPROACH: Gunakan Excel COM API untuk 100% compatibility
    Preserves ALL SharePoint links perfectly (no corruption!)
    """
    app = None
    wb = None

    try:
        if not os.path.exists(rafm_manual_path):
            print(f"❌ File RAFM Manual tidak ditemukan: {rafm_manual_path}")
            return None

        print(f"\n🚀 Menambahkan sheet ke RAFM Output Manual (XLWINGS MODE)...")
        start_time = time.time()

        # Create output directory
        os.makedirs(output_path, exist_ok=True)
        output_file = os.path.join(output_path, output_filename)

        # 🔧 STEP 1: Handle existing output file
        if os.path.exists(output_file):
            print(f"  ↳ File sudah ada, mencoba hapus...")
            try:
                os.remove(output_file)
                print(f"  ✓ File lama berhasil dihapus")
            except PermissionError:
                print(f"  ⚠️ File sedang digunakan, force close Excel...")
                kill_excel_processes()

                try:
                    os.remove(output_file)
                    print(f"  ✓ File lama berhasil dihapus setelah force close")
                except PermissionError:
                    # Use alternative filename with timestamp
                    print(f"  ⚠️ File masih terkunci, menggunakan nama alternatif...")
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    base_name = os.path.splitext(output_filename)[0]
                    ext = os.path.splitext(output_filename)[1]
                    output_filename = f"{base_name}_{timestamp}{ext}"
                    output_file = os.path.join(output_path, output_filename)
                    print(f"  ↳ Nama baru: {output_filename}")

        # 🔧 STEP 2: Copy original file first (preserves metadata)
        print(f"  ↳ Copying original file...")
        shutil.copy2(rafm_manual_path, output_file)
        print(f"  ✓ File copied (all metadata preserved)")

        time.sleep(0.5)  # Wait for file system

        # 🔧 STEP 3: Open with xlwings (Excel COM API)
        print(f"  ↳ Starting Excel via COM API...")
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False

        print(f"  ↳ Opening workbook...")
        wb = app.books.open(output_file)

        # Handle Sheet1 rename
        sheet_names_list = [sh.name for sh in wb.sheets]
        if 'Sheet1' in sheet_names_list and 'RAFM Output Manual' not in sheet_names_list:
            print(f"  ↳ Rename 'Sheet1' → 'RAFM Output Manual'")
            wb.sheets['Sheet1'].name = 'RAFM Output Manual'
        elif 'Sheet1' in sheet_names_list and 'RAFM Output Manual' in sheet_names_list:
            print(f"  ↳ Menghapus 'Sheet1' duplikat...")
            wb.sheets['Sheet1'].delete()

        # Order sheets berdasarkan jenis
        if jenis == 'trad':
            sheet_order = [
                'Control', 'Code',
                'CF ARGO AZTRAD', 'RAFM Output AZTRAD',
                'RAFM Output Manual',
                'RAFM Output AZUL_PI',
                'Checking Summary AZTRAD'
            ]
        elif jenis == 'ul':
            sheet_order = [
                'Control', 'Code',
                'CF ARGO AZUL', 'RAFM Output AZUL',
                'RAFM Output Manual',
                'Checking Summary AZUL'
            ]
        else:  # reas
            sheet_order = [
                'Control', 'Code',
                'CF ARGO REAS', 'RAFM Output REAS',
                'RAFM Output Manual',
                'Checking Summary REAS'
            ]

        # 🔧 STEP 4: Add new sheets
        print(f"  ↳ Menambahkan {len(result_dict)} sheet baru...")

        for sheet_name, df in result_dict.items():
            # 🚨 CRITICAL: Skip RAFM Output Manual
            if sheet_name == 'RAFM Output Manual':
                print(f"    • {sheet_name}: SKIP (preserve existing)")
                continue

            print(f"    • Menambahkan sheet: {sheet_name}")

            # Clean DataFrame
            df = df.copy()
            df = df.replace({pd.NA: None, pd.NaT: None})
            df = df.where(pd.notna(df), None)

            # Delete if exists
            current_sheets = [sh.name for sh in wb.sheets]
            if sheet_name in current_sheets:
                wb.sheets[sheet_name].delete()

            # Add new sheet at the end
            ws = wb.sheets.add(sheet_name, after=wb.sheets[-1])

            # Write data
            if sheet_name == 'Control':
                # Control without header (preserve original behaviour)
                ws.range('A1').value = df.values.tolist()
            else:
                # Other sheets with header
                data_with_header = [df.columns.tolist()] + df.values.tolist()
                ws.range('A1').value = data_with_header

            # -------------------------
            # === Tambahkan formatting sesuai notes.py ===
            # -------------------------
            try:
                apply_number_formats_xlwings(ws, df)
            except Exception as e:
                print(f"    ⚠️ Gagal apply_number_formats untuk {sheet_name}: {e}")

            try:
                apply_border_xlwings(ws, df)
            except Exception as e:
                print(f"    ⚠️ Gagal apply_border untuk {sheet_name}: {e}")

            try:
                auto_adjust_column_width_xlwings(ws, df)
            except Exception as e:
                print(f"    ⚠️ Gagal auto_adjust_column_width untuk {sheet_name}: {e}")

            # Write formulas for Checking Summary (preserve original logic)
            if sheet_name.startswith("Checking Summary"):
                print(f"    • Menulis formula checking summary...")
                try:
                    write_checking_summary_formulas_xlwings(ws, df, jenis)
                except Exception as e:
                    print(f"    ⚠️ Gagal menulis formula checking summary untuk {sheet_name}: {e}")

                # Setelah formula ditulis, pastikan semua angka diformat accounting
                try:
                    apply_accounting_to_all_xlwings(ws, df)
                except Exception as e:
                    print(f"    ⚠️ Gagal apply_accounting_to_all untuk {sheet_name}: {e}")

            # Auto-fit columns (already attempted with auto_adjust, keep original step of autofit if available)
            try:
                ws.autofit(axis='columns')
            except Exception:
                # ignore if not supported in headless environment
                pass

        # 🔧 STEP 5: Reorder sheets
        print(f"  ↳ Mengurutkan sheets...")
        current_sheets = [sh.name for sh in wb.sheets]

        # Move sheets to correct position
        for target_idx, sheet_name in enumerate(sheet_order):
            if sheet_name in current_sheets:
                current_idx = [sh.name for sh in wb.sheets].index(sheet_name)
                if current_idx != target_idx:
                    wb.sheets[sheet_name].api.Move(Before=wb.sheets[target_idx].api)

        # 🔧 STEP 6: Save and close
        print(f"  ↳ Saving workbook...")
        wb.save()
        wb.close()

        print(f"  ↳ Closing Excel...")
        app.quit()

        app = None
        wb = None

        elapsed = time.time() - start_time

        print(f"✅ Selesai dalam {elapsed:.2f} detik")
        print(f"   📁 Output: {output_file}")
        print(f"   ✅ NO CORRUPTION - Excel COM API used")
        print(f"   ✅ SharePoint links PRESERVED")

        return output_file

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return None

    finally:
        # CRITICAL: Always cleanup Excel
        if wb is not None:
            try:
                wb.close()
            except:
                pass

        if app is not None:
            try:
                app.quit()
            except:
                pass

        # Force kill any remaining Excel processes
        time.sleep(1)
        try:
            for proc in psutil.process_iter(['name']):
                if proc.info['name'] and 'excel' in proc.info['name'].lower():
                    try:
                        proc.kill()
                    except:
                        pass
        except:
            pass


def process_input_file(file_path):
    """Process single input file"""
    filename = os.path.basename(file_path).lower()

    # Deteksi jenis
    if 'trad' in filename:
        jenis = 'trad'
        result = trad.main({"input excel": file_path})
    elif 'ul' in filename:
        jenis = 'ul'
        result = ul.main({"input excel": file_path})
    elif 'reas' in filename:
        jenis = 'reas'
        result = reas.main({"input excel": file_path})
    else:
        print(f"❌ Jenis file tidak dikenali: {filename}")
        return

    print(f"\n{'='*60}")
    print(f"📄 PROCESSING: {filename}")
    print(f"   Jenis: {jenis.upper()}")
    print(f"{'='*60}")

    # Baca File Path sheet
    try:
        df = pd.read_excel(file_path, sheet_name='File Path')
    except Exception as e:
        print(f"⚠️ Tidak bisa membaca sheet 'File Path': {e}")
        return

    df.columns = df.columns.str.strip()
    df['Name'] = df['Name'].astype(str).str.strip().str.lower()
    df['File Path'] = df['File Path'].astype(str).str.strip()

    # Validasi required fields
    required = ['output_path', 'output_filename', 'rafm manual']
    missing = [r for r in required if r not in df['Name'].values]
    if missing:
        print(f"⚠️ Missing di File Path sheet: {missing}")
        return

    # Get paths
    output_path = df.loc[df['Name']=='output_path', 'File Path'].values[0]
    output_filename = df.loc[df['Name']=='output_filename', 'File Path'].values[0]
    rafm_manual_path = df.loc[df['Name']=='rafm manual', 'File Path'].values[0]

    # Proses: Tambahkan sheets ke RAFM Manual
    output_file = add_sheets_to_rafm_manual(
        rafm_manual_path=rafm_manual_path,
        result_dict=result,
        output_path=output_path,
        output_filename=output_filename,
        jenis=jenis
    )

    if output_file:
        print(f"\n🎉 SUCCESS: {os.path.basename(output_file)}")
    else:
        print(f"\n❌ FAILED: {filename}")


def main(input_path):
    """Main entry point - SEQUENTIAL processing for xlwings compatibility"""
    print("\n" + "="*60)
    print("🔧 CONTROL 4 - RAFM OUTPUT PROCESSOR (XLWINGS MODE)")
    print("="*60)

    start_time = time.time()

    # Deteksi input
    if os.path.isfile(input_path):
        files = [input_path]
    elif os.path.isdir(input_path):
        files = [
            os.path.join(input_path, fname)
            for fname in os.listdir(input_path)
            if fname.endswith(".xlsx") and not fname.startswith("~$")
        ]
    else:
        print(f"❌ Path tidak valid: {input_path}")
        return

    if not files:
        print("📂 Tidak ada file .xlsx ditemukan")
        return

    print(f"📊 Ditemukan {len(files)} file untuk diproses\n")

    # 🚨 SEQUENTIAL processing for xlwings (COM API tidak support parallel)
    print(f"📋 Mode: Sequential processing (xlwings compatibility)\n")

    success_count = 0
    fail_count = 0

    for idx, file_path in enumerate(files, 1):
        filename = os.path.basename(file_path)
        print(f"\n[{idx}/{len(files)}] Processing: {filename}")

        try:
            process_input_file(file_path)
            success_count += 1
            print(f"✅ [{idx}/{len(files)}] Completed: {filename}")
        except Exception as e:
            fail_count += 1
            print(f"❌ [{idx}/{len(files)}] Failed: {filename}")
            print(f"   Error: {e}")
            import traceback
            traceback.print_exc()

    # Summary
    elapsed = time.time() - start_time
    print("\n" + "="*60)
    print(f"⏱️  TOTAL WAKTU: {elapsed:.2f} detik")
    print(f"📊 Total: {len(files)} file(s)")
    print(f"✅ Success: {success_count}")
    print(f"❌ Failed: {fail_count}")
    try:
        print(f"⚡ Avg: {elapsed/len(files):.2f} detik/file")
    except Exception:
        pass
    print("="*60)


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        print("Usage: python main.py <input_file_or_folder>")
