#!/usr/bin/env python3
"""
Excel Writing Module for Insurance Control System
Handles writing processed results to Excel templates with proper formatting.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import warnings
warnings.filterwarnings('ignore')

def find_template_sheet(wb):
    """Find template sheet in workbook"""
    template_candidates = ['template', 'Template', 'TEMPLATE', 'Control Template']
    
    for sheet_name in wb.sheetnames:
        if any(candidate.lower() in sheet_name.lower() for candidate in template_candidates):
            return sheet_name
    
    # If no template found, use active sheet
    return wb.active.title

def write_summary_data(sheet, df, start_row, title="Summary"):
    """Write summary data to Excel sheet with formatting"""
    if df is None or df.empty:
        return start_row
    
    # Write title
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    start_row += 1
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.value = value
            
            # Format header row
            if r_idx == start_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Format numeric values
            if isinstance(value, (int, float)) and not pd.isna(value):
                if abs(value) >= 1000:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00'
    
    return start_row + len(df) + 3  # Return next available row with spacing

def write_detail_table(sheet, df, start_row, table_name):
    """Write detailed table data to Excel sheet"""
    if df is None or df.empty:
        return start_row
    
    # Write table name
    name_cell = sheet.cell(row=start_row, column=1, value=f"Detail: {table_name}")
    name_cell.font = Font(bold=True, size=11)
    start_row += 1
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.value = value
            
            # Format header row
            if r_idx == start_row:
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            else:
                cell.font = Font(size=9)
            
            # Format numeric values
            if isinstance(value, (int, float)) and not pd.isna(value):
                if abs(value) >= 1000:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00'
    
    return start_row + len(df) + 3  # Return next available row with spacing

def write_trad_results(sheet, result):
    """Write Traditional product results to Excel sheet"""
    current_row = 2  # Start from row 2
    
    # Main Summary
    current_row = write_summary_data(sheet, result.get('summary_total'), current_row, "TRAD Summary")
    
    # Tabel 2 Summary (CC%)
    if result.get('summary_tabel_2') is not None:
        current_row = write_summary_data(sheet, result.get('summary_tabel_2'), current_row, "CC% Summary")
    
    # Tabel 3 Summary (H_IDR_NO)
    if result.get('summary_tabel_3') is not None:
        current_row = write_summary_data(sheet, result.get('summary_tabel_3'), current_row, "H_IDR_NO Summary")
    
    # Tabel 4 Summary (YR)
    if result.get('summary_tabel_4') is not None:
        current_row = write_summary_data(sheet, result.get('summary_tabel_4'), current_row, "YR Summary")
    
    # Tabel 5 Summary (_C_)
    if result.get('summary_tabel_5') is not None:
        current_row = write_summary_data(sheet, result.get('summary_tabel_5'), current_row, "_C_ Summary")
    
    # Detail Tables
    detail_tables = [
        ('Total', result.get('tabel_total')),
        ('CC%', result.get('tabel_2')),
        ('H_IDR_NO', result.get('tabel_3')),
        ('YR', result.get('tabel_4')),
        ('_C_', result.get('tabel_5'))
    ]
    
    for table_name, table_df in detail_tables:
        if table_df is not None and not table_df.empty:
            current_row = write_detail_table(sheet, table_df, current_row, table_name)

def write_ul_results(sheet, result):
    """Write Unit Linked product results to Excel sheet"""
    current_row = 2  # Start from row 2
    
    # Main Summary
    current_row = write_summary_data(sheet, result.get('summary_total'), current_row, "UL Summary")
    
    # Tabel 2 Summary (AG_IDR_SH)
    if result.get('summary_tabel_2') is not None:
        current_row = write_summary_data(sheet, result.get('summary_tabel_2'), current_row, "AG_IDR_SH Summary")
    
    # Tabel 3 Summary (GS)
    if result.get('summary_tabel_3') is not None:
        current_row = write_summary_data(sheet, result.get('summary_tabel_3'), current_row, "GS Summary")
    
    # Detail Tables
    detail_tables = [
        ('Total', result.get('tabel_total')),
        ('AG_IDR_SH', result.get('tabel_2')),
        ('GS', result.get('tabel_3'))
    ]
    
    for table_name, table_df in detail_tables:
        if table_df is not None and not table_df.empty:
            current_row = write_detail_table(sheet, table_df, current_row, table_name)

def create_result_sheet(wb, run_name, result, template_sheet_name):
    """Create a new sheet for each run result"""
    try:
        # Copy template sheet
        template_sheet = wb[template_sheet_name]
        new_sheet = wb.copy_worksheet(template_sheet)
        
        # Create unique sheet name (Excel limit: 31 characters)
        sheet_name = str(run_name)[:31]
        counter = 1
        original_name = sheet_name
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name[:28]}_{counter}"
            counter += 1
        
        new_sheet.title = sheet_name
        
        # Write results based on product type
        product_type = result.get('product_type', 'UNKNOWN')
        if product_type == 'TRAD':
            write_trad_results(new_sheet, result)
        elif product_type == 'UL':
            write_ul_results(new_sheet, result)
        else:
            # Generic writing for unknown types
            write_summary_data(new_sheet, result.get('summary_total'), 2, "Summary")
        
        print(f"Created sheet: {sheet_name} for {product_type} run")
        return True
        
    except Exception as e:
        print(f"Error creating sheet for run {run_name}: {str(e)}")
        return False

def write_to_excel_template(input_path, results_dict):
    """Write all results to Excel template with proper error handling"""
    try:
        # Create backup of original file
        backup_path = input_path.replace('.xlsx', '_backup.xlsx')
        if os.path.exists(input_path):
            import shutil
            shutil.copy2(input_path, backup_path)
            print(f"Backup created: {backup_path}")
        
        # Load workbook
        wb = load_workbook(input_path, keep_vba=True)
        
        # Find template sheet
        template_sheet_name = find_template_sheet(wb)
        print(f"Using template sheet: {template_sheet_name}")
        
        # Process each result
        successful_sheets = 0
        failed_sheets = 0
        
        for run_name, result in results_dict.items():
            if "error" in result:
                print(f"Skipping run {run_name} due to error: {result['error']}")
                failed_sheets += 1
                continue
            
            if create_result_sheet(wb, run_name, result, template_sheet_name):
                successful_sheets += 1
            else:
                failed_sheets += 1
        
        # Save the workbook
        wb.save(input_path)
        print(f"Results saved to: {input_path}")
        print(f"Summary: {successful_sheets} sheets created, {failed_sheets} failed")
        
        return True
        
    except Exception as e:
        print(f"Critical error writing to Excel: {str(e)}")
        return False

def process_and_write_results(input_excel_path, results_dict):
    """Main function to process and write results to Excel"""
    try:
        print("="*60)
        print("Writing Results to Excel")
        print("="*60)
        
        success = write_to_excel_template(input_excel_path, results_dict)
        
        if success:
            print("Excel writing completed successfully!")
        else:
            print("Excel writing failed!")
            
        return success
        
    except Exception as e:
        print(f"Error in process_and_write_results: {str(e)}")
        return False

# Helper functions for formatting
def format_currency(value):
    """Format currency values"""
    if pd.isna(value) or value == 0:
        return 0
    return f"{value:,.2f}"

def format_number(value):
    """Format number values"""
    if pd.isna(value) or value == 0:
        return 0
    return f"{value:,.0f}"

def auto_adjust_column_width(sheet):
    """Auto-adjust column widths based on content"""
    try:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            sheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"Warning: Could not auto-adjust column widths: {str(e)}")

# Additional utility functions that can be used by run_program.py
def validate_excel_file(file_path):
    """Validate that Excel file exists and has required sheets"""
    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"
    
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        
        required_sheets = ['FILTER_TRAD', 'FILTER_UL']
        missing_sheets = [sheet for sheet in required_sheets if sheet not in sheets]
        
        if missing_sheets:
            return False, f"Missing required sheets: {missing_sheets}"
        
        return True, "File validation successful"
        
    except Exception as e:
        return False, f"Error reading Excel file: {str(e)}"

def get_available_sheets(file_path):
    """Get list of available sheets in Excel file"""
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        print(f"Error reading sheets: {str(e)}")
        return []

# This module is designed to be imported by run_program.py
# Main functions to be called:
# - write_to_excel_template(input_path, results_dict)
# - process_and_write_results(input_excel_path, results_dict)
# - validate_excel_file(file_path)