import pandas as pd
import re
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

def parse_multi_values(value):
    """Parse comma/slash separated values"""
    if pd.isna(value) or not value:
        return []
    parts = re.split(r'[,/]', str(value))
    return [p.strip() for p in parts if p.strip()]

def combine_filters(*args):
    """Combine multiple filter lists"""
    combined = []
    for arg in args:
        combined.extend(arg)
    return combined

def apply_filters(df, params):
    """Apply filters to dataframe based on parameters"""
    if df.empty:
        return df.copy()
    
    produk_tertentu = combine_filters(
        parse_multi_values(params.get('only_channel', '')),
        parse_multi_values(params.get('only_currency', '')),
        parse_multi_values(params.get('only_portfolio', '')),
    )
    kecuali_produk = combine_filters(
        parse_multi_values(params.get('exclude_channel', '')),
        parse_multi_values(params.get('exclude_currency', '')),
        parse_multi_values(params.get('exclude_portfolio', '')),
    )
    only_cohort_list = parse_multi_values(params.get('only_cohort', ''))
    only_period_list = parse_multi_values(params.get('only_period', ''))
    
    tahun_tertentu = []
    if only_cohort_list and only_period_list:
        for c in only_cohort_list:
            for p in only_period_list:
                tahun_tertentu.append(f"{c}_{p}")
    
    exclude_cohort_list = parse_multi_values(params.get('exclude_cohort', ''))
    exclude_period_list = parse_multi_values(params.get('exclude_period', ''))
    
    kecuali_tahun = []
    if exclude_cohort_list and exclude_period_list:
        for c in exclude_cohort_list:
            for p in exclude_period_list:
                kecuali_tahun.append(f"{c}_{p}")

    mask = pd.Series(True, index=df.index)

    if kecuali_tahun:
        pattern_exc = '|'.join(map(re.escape, kecuali_tahun))
        mask &= ~df['goc'].astype(str).str.contains(pattern_exc, case=False, na=False)
    
    if tahun_tertentu:
        pattern_inc = '|'.join(map(re.escape, tahun_tertentu))
        mask &= df['goc'].astype(str).str.contains(pattern_inc, case=False, na=False)
    
    if produk_tertentu:
        produk_mask = pd.Series(False, index=df.index)
        for produk in produk_tertentu:
            produk_mask |= df['goc'].astype(str).str.contains(re.escape(produk), case=False, na=False)
        mask &= produk_mask
    
    if kecuali_produk:
        for produk_exc in kecuali_produk:
            mask &= ~df['goc'].astype(str).str.contains(re.escape(produk_exc), case=False, na=False)

    return df[mask].copy()

def filter_goc_by_code(df, code):
    """Filter dataframe by GOC code"""
    if df.empty:
        return df
    tokens = [k for k in code.split('_') if k]
    mask = df['goc'].apply(lambda x: all(token.lower() in str(x).lower() for token in tokens))
    return df[mask].copy()

def exclude_goc_by_code(df, code):
    """Exclude dataframe by GOC code"""
    if df.empty:
        return df
    tokens = [k for k in code.split('_') if k]
    mask = df['goc'].apply(lambda x: all(token.lower() in str(x).lower() for token in tokens))
    return df[~mask].copy()

def clean_numeric_column(df, column_name):
    """Clean and convert column to numeric"""
    if column_name in df.columns:
        df[column_name] = pd.to_numeric(
            df[column_name].astype(str).str.replace(",", ".", regex=False),
            errors="coerce"
        )
        df[column_name] = df[column_name].fillna(0)
    return df

def load_excel_sheet_safely(file_path, sheet_name, required_columns=None):
    """Safely load Excel sheet with error handling"""
    try:
        if not os.path.exists(file_path):
            print(f"Warning: File not found: {file_path}")
            return pd.DataFrame()
        
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        
        if required_columns:
            # Check if required columns exist
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                print(f"Warning: Missing columns {missing_cols} in {sheet_name}")
                return pd.DataFrame()
            df = df[required_columns]
        
        return df
    except Exception as e:
        print(f"Error loading {sheet_name} from {file_path}: {str(e)}")
        return pd.DataFrame()

def run_trad(params):
    """Main function for Traditional products processing"""
    try:
        path_dv = params.get('path_dv', '')
        path_rafm = params.get('path_rafm', '')
        
        if not os.path.isfile(path_dv):
            return {"error": f"File DV tidak ditemukan: {path_dv}"}
        if not os.path.isfile(path_rafm):
            return {"error": f"File RAFM tidak ditemukan: {path_rafm}"}

        # Load DV data
        try:
            dv_trad = pd.read_csv(path_dv)
        except:
            dv_trad = pd.read_excel(path_dv, engine='openpyxl')
        
        # Apply filters
        dv_trad_total = apply_filters(dv_trad, params)
        
        # Drop unnecessary columns
        columns_to_drop = ['product_group', 'pre_ann', 'loan_sa']
        dv_trad_total = dv_trad_total.drop(columns=columns_to_drop, errors='ignore')

        # Process GOC column
        def get_sortir(params):
            def sortir(name):
                if not isinstance(name, str) or not name:
                    return ''
                if '____' in name:
                    double_underscore_parts = name.split('____')
                    if len(double_underscore_parts) > 1:
                        after_double = double_underscore_parts[-1]
                        after_parts = [p for p in after_double.split('_') if p]
                        year_index_after = -1
                        for i, part in enumerate(after_parts):
                            if re.fullmatch(r'\d{4}', part):
                                year_index_after = i
                                break
                        # Check if Q1 in filters
                        only_cohort = parse_multi_values(params.get('only_cohort', ''))
                        only_period = parse_multi_values(params.get('only_period', ''))
                        tahun_tertentu = []
                        for c in only_cohort:
                            for p in only_period:
                                tahun_tertentu.append(f"{c}_{p}")
                        
                        if tahun_tertentu and any('Q1' in t.upper() for t in tahun_tertentu):
                            return after_double
                        if year_index_after == -1:
                            return ''
                        return '_'.join(after_parts[:year_index_after + 1])
                
                parts = [p for p in name.split('_') if p]
                year_index = -1
                for i, part in enumerate(parts):
                    if re.fullmatch(r'\d{4}', part):
                        year_index = i
                        break
                start_index = next((i for i, part in enumerate(parts) if part == 'AG'), 2)
                
                # Check if Q1 in filters
                only_cohort = parse_multi_values(params.get('only_cohort', ''))
                only_period = parse_multi_values(params.get('only_period', ''))
                tahun_tertentu = []
                for c in only_cohort:
                    for p in only_period:
                        tahun_tertentu.append(f"{c}_{p}")
                
                if tahun_tertentu and any('Q1' in t.upper() for t in tahun_tertentu):
                    return '_'.join(parts[start_index:])
                if year_index == -1:
                    return ''
                return '_'.join(parts[start_index:year_index + 1])
            return sortir

        sortir_func = get_sortir(params)
        dv_trad_total['goc'] = dv_trad_total['goc'].apply(sortir_func)
        dv_trad_total['goc'] = dv_trad_total['goc'].apply(lambda x: 'H_IDR_NO_2025' if x == 'IDR_NO_2025' else x)

        # Clean numeric columns
        dv_trad_total = clean_numeric_column(dv_trad_total, 'pol_num')
        dv_trad_total = clean_numeric_column(dv_trad_total, 'sum_assd')

        # Group by GOC
        dv_trad_total = dv_trad_total.groupby(["goc"], as_index=False).sum(numeric_only=True)

        # Apply USD conversion
        usd_rate = float(params.get('USDIDR', 1.0))
        usd_mask = dv_trad_total["goc"].astype(str).str.contains("USD", case=False, na=False)
        dv_trad_total.loc[usd_mask, 'sum_assd'] = dv_trad_total.loc[usd_mask, 'sum_assd'] * usd_rate

        # Load RAFM data
        run_rafm_idr = load_excel_sheet_safely(path_rafm, 'extraction_IDR', ['GOC', 'period', 'cov_units', 'pol_b'])
        run_rafm_usd = load_excel_sheet_safely(path_rafm, 'extraction_USD', ['GOC', 'period', 'cov_units', 'pol_b'])
        
        # Filter period = 0
        if not run_rafm_idr.empty:
            run_rafm_idr = run_rafm_idr[run_rafm_idr['period'].astype(str) == '0']
            run_rafm_idr = run_rafm_idr.drop(columns=["period"])
        
        if not run_rafm_usd.empty:
            run_rafm_usd = run_rafm_usd[run_rafm_usd['period'].astype(str) == '0']
            run_rafm_usd = run_rafm_usd.drop(columns=["period"])

        # Combine RAFM data
        run_rafm_only = pd.concat([run_rafm_idr, run_rafm_usd], ignore_index=True)
        
        if not run_rafm_only.empty:
            run_rafm_only = clean_numeric_column(run_rafm_only, 'pol_b')
            run_rafm_only = clean_numeric_column(run_rafm_only, 'cov_units')
            run_rafm = run_rafm_only.rename(columns={'GOC': 'goc'})
            merged = pd.merge(dv_trad_total, run_rafm, on="goc", how="outer", suffixes=("_trad", "_rafm"))
        else:
            merged = dv_trad_total.copy()
            merged['pol_b'] = 0
            merged['cov_units'] = 0

        merged.fillna(0, inplace=True)
        merged['diff policies'] = merged['pol_num'] - merged['pol_b']
        merged['diff sa'] = merged['sum_assd'] - merged['cov_units']

        # Generate tables
        tabel_total_l = filter_goc_by_code(merged, 'l')
        tabel_total_l = tabel_total_l[~tabel_total_l['goc'].astype(str).str.contains("%", case=False, na=False)]

        # Summary
        summary = pd.DataFrame({
            '': ['Total Trad All from DV', 'Grand Total Summary', 'Check'],
            'DV # of Policies': [
                dv_trad_total['pol_num'].sum(),
                tabel_total_l['pol_num'].sum(),
                dv_trad_total['pol_num'].sum() - tabel_total_l['pol_num'].sum()
            ],
            'DV SA': [
                dv_trad_total['sum_assd'].sum(),
                tabel_total_l['sum_assd'].sum(),
                dv_trad_total['sum_assd'].sum() - tabel_total_l['sum_assd'].sum()
            ],
            'RAFM # of Policies': [
                merged['pol_b'].sum(),
                tabel_total_l['pol_b'].sum(),
                merged['pol_b'].sum() - tabel_total_l['pol_b'].sum()
            ],
            'RAFM SA': [
                merged['cov_units'].sum(),
                tabel_total_l['cov_units'].sum(),
                merged['cov_units'].sum() - tabel_total_l['cov_units'].sum()
            ],
            'Diff # of Policies': [
                dv_trad_total['pol_num'].sum() - merged['pol_b'].sum(),
                tabel_total_l['pol_num'].sum() - tabel_total_l['pol_b'].sum(),
                (dv_trad_total['pol_num'].sum() - merged['pol_b'].sum()) - 
                (tabel_total_l['pol_num'].sum() - tabel_total_l['pol_b'].sum())
            ],
            'Diff SA': [
                dv_trad_total['sum_assd'].sum() - merged['cov_units'].sum(),
                tabel_total_l['sum_assd'].sum() - tabel_total_l['cov_units'].sum(),
                (dv_trad_total['sum_assd'].sum() - merged['cov_units'].sum()) - 
                (tabel_total_l['sum_assd'].sum() - tabel_total_l['cov_units'].sum())
            ]
        })

        # Initialize tables
        tabel_2 = pd.DataFrame()
        tabel_3 = pd.DataFrame()
        tabel_4 = pd.DataFrame()
        tabel_5 = pd.DataFrame()
        summary_tabel_2 = None
        summary_tabel_3 = None
        summary_tabel_4 = None
        summary_tabel_5 = None

        # TABEL 2: CC%
        if params.get('tabel_2_aktif', True):
            tabel_2 = filter_goc_by_code(merged, 'CC%')
            if not tabel_2.empty:
                summary_tabel_2 = pd.DataFrame([{
                    "DV": tabel_2['pol_num'].sum(),
                    "DV SA": tabel_2['sum_assd'].sum(),
                    "RAFM Output": tabel_2['pol_b'].sum(),
                    "RAFM SA": tabel_2['cov_units'].sum(),
                    "Diff # of Policies": tabel_2['pol_num'].sum() - tabel_2['pol_b'].sum(),
                    "Diff SA": tabel_2['sum_assd'].sum() - tabel_2['cov_units'].sum()
                }])

        # TABEL 3: H_IDR_NO
        if params.get('tabel_3_aktif', True):
            tabel_3 = filter_goc_by_code(merged, 'H_IDR_NO')
            if not tabel_3.empty:
                tabel_3_processed = tabel_3.copy()
                tabel_3_processed['goc'] = tabel_3_processed['goc'].apply(
                    lambda x: '_'.join(str(x).split('_')[0:4]) if str(x).startswith('H_IDR_NO') else x
                )
                tabel_3_processed = tabel_3_processed.groupby(['goc'], as_index=False).sum(numeric_only=True)
                
                summary_tabel_3 = pd.DataFrame([{
                    "DV": tabel_3_processed['pol_num'].sum(),
                    "DV SA": tabel_3_processed['sum_assd'].sum(),
                    "RAFM Output": tabel_3_processed['pol_b'].sum(),
                    "RAFM SA": tabel_3_processed['cov_units'].sum(),
                    "Diff # of Policies": tabel_3_processed['pol_num'].sum() - tabel_3_processed['pol_b'].sum(),
                    "Diff SA": tabel_3_processed['sum_assd'].sum() - tabel_3_processed['cov_units'].sum()
                }])
                tabel_3 = tabel_3_processed

        # TABEL 4: YR
        if params.get('tabel_4_aktif', True):
            tabel_4 = filter_goc_by_code(merged, 'YR')
            if not tabel_4.empty:
                tabel_4_processed = tabel_4.copy()
                tabel_4_processed['goc'] = tabel_4_processed['goc'].apply(
                    lambda x: '_'.join(str(x).split('_')[1:5])
                )
                tabel_4_processed = tabel_4_processed.groupby(['goc'], as_index=False).sum(numeric_only=True)
                
                summary_tabel_4 = pd.DataFrame([{
                    "DV": tabel_4_processed['pol_num'].sum(),
                    "DV SA": tabel_4_processed['sum_assd'].sum(),
                    "RAFM Output": tabel_4_processed['pol_b'].sum(),
                    "RAFM SA": tabel_4_processed['cov_units'].sum(),
                    "Diff # of Policies": tabel_4_processed['pol_num'].sum() - tabel_4_processed['pol_b'].sum(),
                    "Diff SA": tabel_4_processed['sum_assd'].sum() - tabel_4_processed['cov_units'].sum()
                }])
                tabel_4 = tabel_4_processed

        # TABEL 5: _C_
        if params.get('tabel_5_aktif', True):
            tabel_5 = filter_goc_by_code(merged, '_C_')
            if not tabel_5.empty:
                tabel_5_processed = tabel_5.copy()
                tabel_5_processed['goc'] = tabel_5_processed['goc'].apply(
                    lambda x: '_'.join(str(x).split('_')[1:5])
                )
                tabel_5_processed = tabel_5_processed.groupby(['goc'], as_index=False).sum(numeric_only=True)
                
                summary_tabel_5 = pd.DataFrame([{
                    "DV": tabel_5_processed['pol_num'].sum(),
                    "DV SA": tabel_5_processed['sum_assd'].sum(),
                    "RAFM Output": tabel_5_processed['pol_b'].sum(),
                    "RAFM SA": tabel_5_processed['cov_units'].sum(),
                    "Diff # of Policies": tabel_5_processed['pol_num'].sum() - tabel_5_processed['pol_b'].sum(),
                    "Diff SA": tabel_5_processed['sum_assd'].sum() - tabel_5_processed['cov_units'].sum()
                }])
                tabel_5 = tabel_5_processed

        return {
            'tabel_total': tabel_total_l,
            'tabel_2': tabel_2,
            'tabel_3': tabel_3,
            'tabel_4': tabel_4,
            'tabel_5': tabel_5,
            'summary_total': summary,
            'summary_tabel_2': summary_tabel_2,
            'summary_tabel_3': summary_tabel_3,
            'summary_tabel_4': summary_tabel_4,
            'summary_tabel_5': summary_tabel_5,
            'run_name': params.get('run_name', '')
        }

    except Exception as e:
        return {"error": f"Error in run_trad: {str(e)}"}

def run_ul(params):
    """Main function for Unit Linked products processing"""
    try:
        path_dv = params.get('path_dv', '')
        path_rafm = params.get('path_rafm', '')
        path_uvsg = params.get('path_uvsg', '')
        
        if not os.path.isfile(path_dv):
            return {"error": f"File DV tidak ditemukan: {path_dv}"}
        if not os.path.isfile(path_rafm):
            return {"error": f"File RAFM tidak ditemukan: {path_rafm}"}

        # Load DV data
        dv_ul = load_excel_sheet_safely(path_dv, 0)  # Use first sheet
        if dv_ul.empty:
            return {"error": "Failed to load DV data"}
        
        dv_ul_total = apply_filters(dv_ul, params)
        dv_ul_total = dv_ul_total.drop(columns=['product_group', 'pre_ann', 'sum_assur'], errors='ignore')

        # Process GOC
        def sortir(name):
            parts = [p for p in str(name).split('_') if p]
            year_index = -1
            for i, part in enumerate(parts):
                if re.fullmatch(r'\d{4}', part):
                    year_index = i
                    break
            if year_index == -1:
                return ''
            start_index = None
            for i, part in enumerate(parts):
                if part == 'AG':
                    start_index = i
                    break
            if start_index is None:
                start_index = 2
            return '_'.join(parts[start_index:year_index+1])

        dv_ul_total['goc'] = dv_ul_total['goc'].apply(sortir)
        dv_ul_total = clean_numeric_column(dv_ul_total, 'total_fund')
        dv_ul_total = dv_ul_total.groupby(["goc"], as_index=False).sum(numeric_only=True)

        # Apply USD conversion
        usd_rate = float(params.get('USDIDR', 1.0))
        usd_mask = dv_ul_total["goc"].astype(str).str.contains("USD", case=False, na=False)
        dv_ul_total.loc[usd_mask, 'total_fund'] = dv_ul_total.loc[usd_mask, 'total_fund'] * usd_rate

        # Load RAFM data
        run_rafm_idr = load_excel_sheet_safely(path_rafm, 'extraction_IDR', ['GOC', 'period', 'pol_b', 'RV_AV_IF'])
        run_rafm_usd = load_excel_sheet_safely(path_rafm, 'extraction_USD', ['GOC', 'period', 'pol_b', 'RV_AV_IF'])
        
        # Filter period = 0
        if not run_rafm_idr.empty:
            run_rafm_idr = run_rafm_idr[run_rafm_idr['period'].astype(str) == '0']
            run_rafm_idr = run_rafm_idr.drop(columns=["period"])
        
        if not run_rafm_usd.empty:
            run_rafm_usd = run_rafm_usd[run_rafm_usd['period'].astype(str) == '0']
            run_rafm_usd = run_rafm_usd.drop(columns=["period"])

        # Combine RAFM data
        run_rafm_only = pd.concat([run_rafm_idr, run_rafm_usd], ignore_index=True)
        if not run_rafm_only.empty:
            run_rafm_only = clean_numeric_column(run_rafm_only, 'pol_b')
            run_rafm_only = clean_numeric_column(run_rafm_only, 'RV_AV_IF')

        # Exclude GS from RAFM
        run_rafm_no_gs = run_rafm_only[~run_rafm_only['GOC'].astype(str).str.contains('GS', case=False, na=False)]

        # Load UVSG data if provided
        run_uvsg = pd.DataFrame()
        if path_uvsg and os.path.isfile(path_uvsg):
            run_uvsg_idr = load_excel_sheet_safely(path_uvsg, 'extraction_IDR', ['GOC', 'period', 'pol_b', 'rv_av_if'])
            run_uvsg_usd = load_excel_sheet_safely(path_uvsg, 'extraction_USD', ['GOC', 'period', 'pol_b', 'rv_av_if'])
            
            if not run_uvsg_idr.empty:
                run_uvsg_idr = run_uvsg_idr[run_uvsg_idr['period'].astype(str) == '0']
                run_uvsg_idr = run_uvsg_idr.drop(columns=["period"])
            
            if not run_uvsg_usd.empty:
                run_uvsg_usd = run_uvsg_usd[run_uvsg_usd['period'].astype(str) == '0']
                run_uvsg_usd = run_uvsg_usd.drop(columns=["period"])

            run_uvsg = pd.concat([run_uvsg_idr, run_uvsg_usd], ignore_index=True)
            if not run_uvsg.empty:
                run_uvsg = clean_numeric_column(run_uvsg, 'pol_b')
                run_uvsg = clean_numeric_column(run_uvsg, 'rv_av_if')
                run_uvsg = run_uvsg.rename(columns={'rv_av_if': 'RV_AV_IF'})

        # Combine RAFM and UVSG
        run_rafm = pd.concat([run_rafm_no_gs, run_uvsg], ignore_index=True)
        run_rafm = run_rafm.rename(columns={'GOC': 'goc'})

        # Merge data
        merged = pd.merge(dv_ul_total, run_rafm, on="goc", how="outer", suffixes=("_uv_total", "_run_rafm"))
        merged.fillna(0, inplace=True)

        merged['diff policies'] = merged.get('pol_num', 0) - merged.get('pol_b', 0)
        merged['diff sa'] = merged.get('total_fund', 0) - merged.get('RV_AV_IF', 0)

        # Generate tables
        tabel_total_l = exclude_goc_by_code(merged, 'gs')

        # Summary
        summary = pd.DataFrame({
            '': ['Total UL All from DV', 'Grand Total Summary', 'Check'],
            'DV # of Policies': [
                dv_ul_total['pol_num'].sum() if 'pol_num' in dv_ul_total else 0,
                tabel_total_l['pol_num'].sum() if 'pol_num' in tabel_total_l else 0,
                (dv_ul_total['pol_num'].sum() if 'pol_num' in dv_ul_total else 0) - 
                (tabel_total_l['pol_num'].sum() if 'pol_num' in tabel_total_l else 0)
            ],
            'DV Fund Value': [
                dv_ul_total['total_fund'].sum() if 'total_fund' in dv_ul_total else 0,
                tabel_total_l['total_fund'].sum() if 'total_fund' in tabel_total_l else 0,
                (dv_ul_total['total_fund'].sum() if 'total_fund' in dv_ul_total else 0) - 
                (tabel_total_l['total_fund'].sum() if 'total_fund' in tabel_total_l else 0)
            ],
            'RAFM # of Policies': [
                run_rafm['pol_b'].sum() if 'pol_b' in run_rafm and not run_rafm.empty else 0,
                tabel_total_l['pol_b'].sum() if 'pol_b' in tabel_total_l else 0,
                (run_rafm['pol_b'].sum() if 'pol_b' in run_rafm and not run_rafm.empty else 0) - 
                (tabel_total_l['pol_b'].sum() if 'pol_b' in tabel_total_l else 0)
            ],
            'RAFM Fund Value': [
                run_rafm['RV_AV_IF'].sum() if 'RV_AV_IF' in run_rafm and not run_rafm.empty else 0,
                tabel_total_l['RV_AV_IF'].sum() if 'RV_AV_IF' in tabel_total_l else 0,
                (run_rafm['RV_AV_IF'].sum() if 'RV_AV_IF' in run_rafm and not run_rafm.empty else 0) - 
                (tabel_total_l['RV_AV_IF'].sum() if 'RV_AV_IF' in tabel_total_l else 0)
            ],
            'Diff # of Policies': [
                (dv_ul_total['pol_num'].sum() if 'pol_num' in dv_ul_total else 0) - 
                (run_rafm['pol_b'].sum() if 'pol_b' in run_rafm and not run_rafm.empty else 0),
                (tabel_total_l['pol_num'].sum() if 'pol_num' in tabel_total_l else 0) - 
                (tabel_total_l['pol_b'].sum() if 'pol_b' in tabel_total_l else 0),
                ((dv_ul_total['pol_num'].sum() if 'pol_num' in dv_ul_total else 0) - 
                 (run_rafm['pol_b'].sum() if 'pol_b' in run_rafm and not run_rafm.empty else 0)) -
                ((tabel_total_l['pol_num'].sum() if 'pol_num' in tabel_total_l else 0) - 
                 (tabel_total_l['pol_b'].sum() if 'pol_b' in tabel_total_l else 0))
            ],
            'Diff Fund Value': [
                (dv_ul_total['total_fund'].sum() if 'total_fund' in dv_ul_total else 0) - 
                (run_rafm['RV_AV_IF'].sum() if 'RV_AV_IF' in run_rafm and not run_rafm.empty else 0),
                (tabel_total_l['total_fund'].sum() if 'total_fund' in tabel_total_l else 0) - 
                (tabel_total_l['RV_AV_IF'].sum() if 'RV_AV_IF' in tabel_total_l else 0),
                ((dv_ul_total['total_fund'].sum() if 'total_fund' in dv_ul_total else 0) - 
                 (run_rafm['RV_AV_IF'].sum() if 'RV_AV_IF' in run_rafm and not run_rafm.empty else 0)) -
                ((tabel_total_l['total_fund'].sum() if 'total_fund' in tabel_total_l else 0) - 
                 (tabel_total_l['RV_AV_IF'].sum() if 'RV_AV_IF' in tabel_total_l else 0))
            ]
        })

        # Initialize tables
        tabel_2 = pd.DataFrame()
        tabel_3 = pd.DataFrame()
        summary_tabel_2 = None
        summary_tabel_3 = None

        # TABEL 2: AG_IDR_SH
        if params.get('tabel_2_aktif', True):
            tabel_2 = filter_goc_by_code(merged, 'AG_IDR_SH')
            if not tabel_2.empty:
                summary_tabel_2 = pd.DataFrame([{
                    "DV": tabel_2['pol_num'].sum() if 'pol_num' in tabel_2 else 0,
                    "DV Fund": tabel_2['total_fund'].sum() if 'total_fund' in tabel_2 else 0,
                    "RAFM Output": tabel_2['pol_b'].sum() if 'pol_b' in tabel_2 else 0,
                    "RAFM Fund": tabel_2['RV_AV_IF'].sum() if 'RV_AV_IF' in tabel_2 else 0,
                    'Diff # of Policies': (tabel_2['pol_num'].sum() if 'pol_num' in tabel_2 else 0) - 
                                        (tabel_2['pol_b'].sum() if 'pol_b' in tabel_2 else 0),
                    'Diff fund': (tabel_2['total_fund'].sum() if 'total_fund' in tabel_2 else 0) - 
                               (tabel_2['RV_AV_IF'].sum() if 'RV_AV_IF' in tabel_2 else 0)
                }])

        # TABEL 3: GS
        if params.get('tabel_3_aktif', True):
            # Get GS data from original RAFM
            tabel_gs = filter_goc_by_code(run_rafm_only, 'GS') if not run_rafm_only.empty else pd.DataFrame()
            tabel_gs = tabel_gs.rename(columns={'GOC': 'goc'}) if not tabel_gs.empty else tabel_gs
            dv_gs = filter_goc_by_code(dv_ul_total, 'GS')

            tabel_3 = pd.merge(dv_gs, tabel_gs, on="goc", how="outer", suffixes=("_uv_total", "_run_rafm"))
            tabel_3.fillna(0, inplace=True)

            tabel_3['diff policies'] = tabel_3.get('pol_num', 0) - tabel_3.get('pol_b', 0)
            tabel_3['diff sa'] = tabel_3.get('total_fund', 0) - tabel_3.get('RV_AV_IF', 0)

            if not tabel_3.empty:
                summary_tabel_3 = pd.DataFrame([{
                    "DV": tabel_3['pol_num'].sum(),
                    "DV Fund": tabel_3['total_fund'].sum(),
                    "RAFM Output": tabel_3['pol_b'].sum(),
                    "RAFM Fund": tabel_3['RV_AV_IF'].sum(),
                    "Diff # of Policies": tabel_3['pol_num'].sum() - tabel_3['pol_b'].sum(),
                    "Diff Fund": tabel_3['total_fund'].sum() - tabel_3['RV_AV_IF'].sum()
                }])

        return {
            'tabel_total': tabel_total_l,
            'tabel_2': tabel_2,
            'tabel_3': tabel_3,
            'summary_total': summary,
            'summary_tabel_2': summary_tabel_2,
            'summary_tabel_3': summary_tabel_3,
            'run_name': params.get('run_name', '')
        }

    except Exception as e:
        return {"error": f"Error in run_ul: {str(e)}"}

def load_filters(sheet_path, sheet_name):
    """Load filter configurations from Excel sheet"""
    try:
        df = pd.read_excel(sheet_path, sheet_name=sheet_name, engine='openpyxl')
        df.columns = df.columns.str.strip()
        
        # Check if RUN column exists, if not use run_name
        if 'RUN' not in df.columns and 'run_name' in df.columns:
            df['RUN'] = df['run_name']
        elif 'RUN' not in df.columns:
            df['RUN'] = df.index + 1  # Create default run numbers
        
        # Filter out rows where RUN is null
        df = df[df['RUN'].notna()]
        
        # Convert to dictionary format expected by the processing functions
        configs = []
        for _, row in df.iterrows():
            config = {
                'RUN': row.get('RUN', ''),
                'run_name': row.get('run_name', row.get('RUN', '')),
                'path_dv': row.get('path_dv', ''),
                'path_rafm': row.get('path_rafm', ''),
                'path_uvsg': row.get('path_uvsg', ''),  # Only for UL
                'USDIDR': float(row.get('USDIDR', 1.0)),
                'only_channel': row.get('only_channel', ''),
                'exclude_channel': row.get('exclude_channel', ''),
                'only_currency': row.get('only_currency', ''),
                'exclude_currency': row.get('exclude_currency', ''),
                'only_portfolio': row.get('only_portfolio', ''),
                'exclude_portfolio': row.get('exclude_portfolio', ''),
                'only_cohort': row.get('only_cohort', ''),
                'exclude_cohort': row.get('exclude_cohort', ''),
                'only_period': row.get('only_period', ''),
                'exclude_period': row.get('exclude_period', ''),
                'tabel_2_aktif': bool(row.get('tabel_2_aktif', True)),
                'tabel_3_aktif': bool(row.get('tabel_3_aktif', True)),
                'tabel_4_aktif': bool(row.get('tabel_4_aktif', True)),  # Only for TRAD
                'tabel_5_aktif': bool(row.get('tabel_5_aktif', True)),  # Only for TRAD
            }
            configs.append(config)
        
        return configs
    except Exception as e:
        print(f"Error loading filters from {sheet_name}: {str(e)}")
        return []

def process_trad_run(run_config):
    """Process Traditional run with error handling"""
    try:
        print(f"Processing TRAD run: {run_config['RUN']}")
        result = run_trad(run_config)
        return run_config["RUN"], result
    except Exception as e:
        print(f"Error in TRAD run {run_config['RUN']}: {str(e)}")
        return run_config["RUN"], {"error": str(e)}

def process_ul_run(run_config):
    """Process Unit Linked run with error handling"""
    try:
        print(f"Processing UL run: {run_config['RUN']}")
        result = run_ul(run_config)
        return run_config["RUN"], result
    except Exception as e:
        print(f"Error in UL run {run_config['RUN']}: {str(e)}")
        return run_config["RUN"], {"error": str(e)}

def write_to_excel_template(input_path, results_dict):
    """Write results to Excel template with proper error handling"""
    try:
        # Create backup of original file
        backup_path = input_path.replace('.xlsx', '_backup.xlsx')
        if os.path.exists(input_path):
            import shutil
            shutil.copy2(input_path, backup_path)
        
        wb = load_workbook(input_path, keep_vba=True)
        
        # Find template sheet
        template_sheet_name = None
        for sheet_name in wb.sheetnames:
            if 'template' in sheet_name.lower():
                template_sheet_name = sheet_name
                break
        
        if template_sheet_name:
            template_sheet = wb[template_sheet_name]
        else:
            # Create a simple template if none exists
            template_sheet = wb.active
            template_sheet.title = "Control Template"

        for run_name, result in results_dict.items():
            if "error" in result:
                print(f"Skipping run {run_name} due to error: {result['error']}")
                continue

            try:
                # Create new sheet for each run
                new_sheet = wb.copy_worksheet(template_sheet)
                
                # Truncate sheet name to 31 characters (Excel limit)
                sheet_name = str(run_name)[:31]
                
                # Ensure unique sheet name
                counter = 1
                original_name = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_name[:28]}_{counter}"
                    counter += 1
                
                new_sheet.title = sheet_name

                # Write Summary (starting at A2)
                if "summary_total" in result and result["summary_total"] is not None:
                    summary_df = result["summary_total"]
                    start_row = 2
                    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            cell = new_sheet.cell(row=r_idx, column=c_idx)
                            cell.value = value

                # Write Tabel 2 Summary (starting at A20)
                if "summary_tabel_2" in result and result["summary_tabel_2"] is not None:
                    df2 = result["summary_tabel_2"]
                    start_row = 20
                    for r_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            cell = new_sheet.cell(row=r_idx, column=c_idx)
                            cell.value = value

                # Write Tabel 3 Summary (starting at A30)
                if "summary_tabel_3" in result and result["summary_tabel_3"] is not None:
                    df3 = result["summary_tabel_3"]
                    start_row = 30
                    for r_idx, row in enumerate(dataframe_to_rows(df3, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            cell = new_sheet.cell(row=r_idx, column=c_idx)
                            cell.value = value

                # Write Tabel 4 Summary (starting at A40) - Only for TRAD
                if "summary_tabel_4" in result and result["summary_tabel_4"] is not None:
                    df4 = result["summary_tabel_4"]
                    start_row = 40
                    for r_idx, row in enumerate(dataframe_to_rows(df4, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            cell = new_sheet.cell(row=r_idx, column=c_idx)
                            cell.value = value

                # Write Tabel 5 Summary (starting at A50) - Only for TRAD
                if "summary_tabel_5" in result and result["summary_tabel_5"] is not None:
                    df5 = result["summary_tabel_5"]
                    start_row = 50
                    for r_idx, row in enumerate(dataframe_to_rows(df5, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            cell = new_sheet.cell(row=r_idx, column=c_idx)
                            cell.value = value

                # Write detailed tables if needed (starting at A60)
                detail_start_row = 60
                for table_name in ['tabel_total', 'tabel_2', 'tabel_3', 'tabel_4', 'tabel_5']:
                    if table_name in result and result[table_name] is not None and not result[table_name].empty:
                        # Add table header
                        new_sheet.cell(row=detail_start_row, column=1, value=f"Detail {table_name}")
                        detail_start_row += 1
                        
                        # Add table data
                        table_df = result[table_name]
                        for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=True), start=detail_start_row):
                            for c_idx, value in enumerate(row, start=1):
                                cell = new_sheet.cell(row=r_idx, column=c_idx)
                                cell.value = value
                        
                        detail_start_row += len(table_df) + 3  # Add some spacing

                print(f"Successfully created sheet: {sheet_name}")

            except Exception as e:
                print(f"Error creating sheet for run {run_name}: {str(e)}")
                continue

        # Save the workbook
        wb.save(input_path)
        print(f"Results saved to: {input_path}")
        print(f"Backup created at: {backup_path}")

    except Exception as e:
        print(f"Error writing to Excel: {str(e)}")
        raise

def main(INPUT_EXCEL_PATH):
    """Main function to orchestrate the entire process"""
    try:
        print(f"Starting processing of: {INPUT_EXCEL_PATH}")
        
        if not os.path.exists(INPUT_EXCEL_PATH):
            raise FileNotFoundError(f"Input file not found: {INPUT_EXCEL_PATH}")

        # Load filter configurations
        print("Loading TRAD filters...")
        trad_filters = load_filters(INPUT_EXCEL_PATH, "FILTER_TRAD")
        print(f"Loaded {len(trad_filters)} TRAD configurations")

        print("Loading UL filters...")
        ul_filters = load_filters(INPUT_EXCEL_PATH, "FILTER_UL")
        print(f"Loaded {len(ul_filters)} UL configurations")

        trad_results = {}
        ul_results = {}

        # Process runs using ThreadPoolExecutor for parallel processing
        print("Starting parallel processing...")
        
        with ThreadPoolExecutor(max_workers=4) as executor:
            # Submit TRAD jobs
            trad_futures = {}
            for cfg in trad_filters:
                future = executor.submit(process_trad_run, cfg)
                trad_futures[future] = cfg["RUN"]

            # Submit UL jobs
            ul_futures = {}
            for cfg in ul_filters:
                future = executor.submit(process_ul_run, cfg)
                ul_futures[future] = cfg["RUN"]

            # Collect TRAD results
            print("Collecting TRAD results...")
            for future in as_completed(trad_futures):
                run_name, result = future.result()
                trad_results[run_name] = result
                print(f"Completed TRAD run: {run_name}")

            # Collect UL results
            print("Collecting UL results...")
            for future in as_completed(ul_futures):
                run_name, result = future.result()
                ul_results[run_name] = result
                print(f"Completed UL run: {run_name}")

        # Combine all results
        all_results = {**trad_results, **ul_results}
        
        print(f"Processing completed. Total runs: {len(all_results)}")
        
        # Write results to Excel
        print("Writing results to Excel...")
        write_to_excel_template(INPUT_EXCEL_PATH, all_results)
        
        print("Process completed successfully!")
        
        # Print summary
        successful_runs = sum(1 for result in all_results.values() if "error" not in result)
        failed_runs = len(all_results) - successful_runs
        print(f"\nSummary:")
        print(f"- Total runs: {len(all_results)}")
        print(f"- Successful: {successful_runs}")
        print(f"- Failed: {failed_runs}")
        
        if failed_runs > 0:
            print("\nFailed runs:")
            for run_name, result in all_results.items():
                if "error" in result:
                    print(f"- {run_name}: {result['error']}")

    except Exception as e:
        print(f"Critical error in main process: {str(e)}")
        raise

# Example usage
if __name__ == "__main__":
    # Update this path to your actual Excel file
    INPUT_SHEET_PATH = r"D:\Run Control 3\IRCS3_build\Input Sheet_IRCS3.xlsx"
    
    # Alternative path for testing - update as needed
    # INPUT_SHEET_PATH = "Input Sheet_IRCS3.xlsx"
    
    try:
        main(INPUT_SHEET_PATH)
    except Exception as e:
        print(f"Program failed: {str(e)}")
        input("Press Enter to exit...")  # Keep console open to see error